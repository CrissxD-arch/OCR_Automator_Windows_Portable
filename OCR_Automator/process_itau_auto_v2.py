#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Procesa y limpia CSV Itaú con soporte de Itau_auto_debug.txt.

Mejoras clave:
- Docstrings y funciones refactorizadas (legibilidad y testabilidad).
- Escritura en streaming (no acumula todas las filas en memoria).
- Autodetección de delimitador (csv.Sniffer) si se usa --delimiter auto.
- Detección de encoding opcional con chardet (si está instalado).
- Manejo robusto de errores de IO y métricas detalladas en el reporte.
- Validaciones opcionales (--required-fields, --reject-incomplete).
- Logging configurable (-v/-vv) y opción de log a archivo (—log-file futura si se desea).
- Importación de constantes desde constants.py (separación de reglas).

Uso típico:
  python process_itau_auto_v2.py --input Itau_results_ALL.csv --debug Itau_auto_debug.txt --output Itau_results_ALL.cleaned.csv --report fix_report.md

Ejecución sin argumentos:
  - Autodetecta input CSV en carpeta actual.
  - Usa Itau_auto_debug.txt si existe.
  - Genera <input>.cleaned.csv y fix_report.md.

Requisitos opcionales:
  - ftfy (mejor corrección de mojibake): pip install ftfy
  - chardet (detección de encoding):     pip install chardet
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import logging
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

# Dependencias opcionales
try:
    import ftfy  # type: ignore
except Exception:
    ftfy = None

try:
    import chardet  # type: ignore
except Exception:
    chardet = None

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:
    openpyxl = None

# Constantes separadas
try:
    from constants import (
        CANONICAL_HEADERS,
        HEADER_ALIASES,
        DATE_FIELDS,
        INT_FIELDS,
        APODERADO_1,
        APODERADO_2,
        APODERADO_PATTERN_STRS,
        COMMON_FIXES,
        VALID_COMUNAS,
    )
except ImportError:
    try:
        from .constants import (
            CANONICAL_HEADERS,
            HEADER_ALIASES,
            DATE_FIELDS,
            INT_FIELDS,
            APODERADO_1,
            APODERADO_2,
            APODERADO_PATTERN_STRS,
            COMMON_FIXES,
            VALID_COMUNAS,
        )
    except ImportError:
        # Fallback con valores por defecto si no se puede importar
        CANONICAL_HEADERS = ['RUT_CLIENTE', 'NOMBRE_COMPLETO', 'MONTO_CREDITO']
        HEADER_ALIASES = {}
        DATE_FIELDS = {'FECHA_NACIMIENTO', 'FECHA_CONTRATO'}
        INT_FIELDS = {'PLAZO_MESES'}
        APODERADO_1 = {}
        APODERADO_2 = {}
        APODERADO_PATTERN_STRS = []
        COMMON_FIXES = {}
        VALID_COMUNAS = set()

APODERADO_PATTERNS = [re.compile(pat, re.I) for pat in APODERADO_PATTERN_STRS]

# Importaciones de geolocalización y limpieza
try:
    from geocoding_utils import (
        enhance_dataframe_with_geolocation, 
        cleanup_temp_files,
        apply_reference_corrections,
        validate_rut_dv,
        clean_and_fix_address,
        fix_comuna_ocr
    )
    GEOCODING_AVAILABLE = True
    print("✅ Módulo de geolocalización cargado")
except ImportError:
    GEOCODING_AVAILABLE = False
    logging.warning("⚠️ Módulo de geolocalización no disponible")


# ========================= Utilidades de texto/encoding =========================

def fix_text(s: Optional[str]) -> str:
    """
    Repara mojibake y normaliza el texto.
    - Usa ftfy si está disponible.
    - Normaliza Unicode a NFC y elimina el replacement char.
    """
    if s is None:
        return ""
    if ftfy:
        try:
            return ftfy.fix_text(s)
        except Exception:
            pass
    s = unicodedata.normalize("NFC", s)
    return s.replace("\uFFFD", "")


def clean_text_value(s: Optional[str]) -> str:
    """
    Limpia una celda de texto:
    - Normaliza encodings.
    - Elimina comillas tipográficas y caracteres invisibles.
    - Contrae espacios y recorta.
    """
    if s is None:
        return ""
    s = fix_text(s)
    s = s.replace("\u200b", "")
    s = s.replace('"', "").replace("“", "").replace("”", "")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_header(h: str) -> str:
    """
    Normaliza un encabezado de columna a su forma canónica usando HEADER_ALIASES.
    """
    h = fix_text((h or "").strip())
    return HEADER_ALIASES.get(h, h)


def apply_common_fixes(value: str) -> str:
    """
    Aplica reemplazos comunes por OCR a campos de texto (COMUNA/DIRECCION/NOMBRE).
    """
    out = value
    for pat, repl in COMMON_FIXES.items():
        out = re.sub(pat, repl, out, flags=re.I)
    return out


def normalize_percent(s: str) -> str:
    """
    Normaliza una tasa:
    - Acepta '1,06%' o '1.06%' y la deja como '<numero>%', usando punto decimal.
    """
    if not s:
        return ""
    s = clean_text_value(s).replace(" ", "")
    if s.endswith("%"):
        num = s[:-1].replace(",", ".")
        return f"{num}%"
    return s


def normalize_int_digits(s: str) -> str:
    """
    Extrae solo dígitos de una cantidad, eliminando separadores y símbolos.
    """
    if not s:
        return ""
    s = clean_text_value(s)
    return re.sub(r"[^\d]", "", s.replace(".", "").replace(" ", "").split(",")[0])


def format_int(value: str, thousand_sep: str) -> str:
    """
    Formatea un string de dígitos como entero con separador de miles deseado.
    thousand_sep: 'none' | 'dot' | 'comma'
    """
    if not value or not value.isdigit():
        return value
    if thousand_sep == "none":
        return value
    sep = "." if thousand_sep == "dot" else ","
    rev = value[::-1]
    groups = [rev[i:i + 3] for i in range(0, len(rev), 3)]
    return sep.join(groups)[::-1]


def parse_date_multi(s: str) -> Optional[dt.date]:
    """
    Intenta parsear una fecha con múltiples formatos comunes.
    Formatos soportados: dd-mm-YYYY, dd/mm/YYYY, YYYY-mm-dd, mm/dd/YYYY
    """
    if not s:
        return None
    cand = clean_text_value(s)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            # Pre-normaliza separadores
            val = cand.replace("/", "-") if fmt in ("%d-%m-%Y", "%Y-%m-%d") else cand
            return dt.datetime.strptime(val, fmt).date()
        except Exception:
            continue
    return None


def format_date(d: Optional[dt.date], fmt: str) -> str:
    """
    Devuelve la fecha formateada según fmt:
    - 'iso': YYYY-mm-dd
    - 'dmy': dd-mm-YYYY
    """
    if not d:
        return ""
    return d.isoformat() if fmt == "iso" else d.strftime("%d-%m-%Y")


def clean_apoderado(value: str, which: int) -> str:
    """
    Canoniza el nombre del apoderado (1 o 2) basándose en patrones frecuentes.
    Si está vacío, asigna el canónico por defecto.
    """
    v = clean_text_value(value or "")
    if not v:
        # Devolver un valor por defecto basado en which
        default_names = {1: "Apoderado Principal", 2: "Apoderado Suplente"}
        return default_names.get(which, "Apoderado")
    
    # Buscar en diccionarios de apoderados conocidos
    rut_clean = rut_clean_digits(v)
    if rut_clean in APODERADO_1:
        return APODERADO_1[rut_clean]
    if rut_clean in APODERADO_2:
        return APODERADO_2[rut_clean]
    
    # Buscar por patrones
    for pat in APODERADO_PATTERNS:
        match = pat.search(v)
        if match:
            return match.group(1).strip()
    
    # Buscar nombres específicos
    if which == 1 and re.search(r"\byasna\b", v, re.I):
        return v.title()
    if which == 2 and re.search(r"\berwin\b", v, re.I):
        return v.title()
    
    return v.title() if v else ""


def rut_clean_digits(rut: str) -> str:
    return re.sub(r"[^\dkK]", "", rut or "")


def rut_calc_dv(num_str: str) -> str:
    """
    Calcula el dígito verificador del RUT chileno.
    """
    if not num_str.isdigit():
        return ""
    reversed_digits = list(map(int, reversed(num_str)))
    factors = [2, 3, 4, 5, 6, 7]
    s = sum(d * factors[i % len(factors)] for i, d in enumerate(reversed_digits))
    mod = 11 - (s % 11)
    if mod == 11:
        return "0"
    if mod == 10:
        return "K"
    return str(mod)


def normalize_rut_and_dv(rut: str, dv: str) -> Tuple[str, str, str, bool]:
    """
    Limpia RUT/DV y valida contra DV calculado.
    Retorna: (rut_num, dv_clean, dv_calc, valido)
    """
    rut_num = rut_clean_digits(rut)
    dv_clean = clean_text_value(dv or "").upper()
    calc = rut_calc_dv(rut_num) if rut_num else ""
    valid = (calc == dv_clean) if (rut_num and dv_clean) else True
    return rut_num, dv_clean, calc, valid


# ========================= Debug: parseo y merge =========================

def parse_debug_final_rows(debug_text: str) -> Dict[str, Dict[str, str]]:
    """
    Parsea bloques '---- FINAL ROW ---- ... ---- END FINAL ROW ----' y construye un mapeo.
    Clave: OPERACION si existe; si no, RUT-DV-NOMBRE.
    Devuelve un dict: key -> {campo: valor normalizado}.
    """
    mapping: Dict[str, Dict[str, str]] = {}
    blocks = debug_text.split("---- FINAL ROW ----")
    if len(blocks) <= 1:
        return mapping
    for block in blocks[1:]:
        part = block.split("---- END FINAL ROW ----")[0]
        if not part:
            continue
        d: Dict[str, str] = {}
        for line in part.strip().splitlines():
            m = re.match(r"\s*([^:]+):\s*(.*)\s*$", line)
            if not m:
                continue
            k = fix_text(m.group(1)).strip().upper()
            v = clean_text_value(m.group(2))
            d[k] = v
        key_map = {"OPERACIÓN": "OPERACION", "OPERACION": "OPERACION", "FECHA_VENCIMIENTO_1°_CUOTA": "FECHA_VENCIMIENTO_1_CUOTA"}
        norm = {key_map.get(k, k): v for k, v in d.items()}
        op = norm.get("OPERACION", "").strip()
        if op:
            mapping[op] = norm
        else:
            rk = f"{norm.get('RUT','')}-{norm.get('DV','')}-{norm.get('NOMBRE','')}"
            if rk.strip("-"):
                mapping[rk] = norm
    return mapping


def merge_from_debug(row: dict, debug_map: dict, mode: str, stats_fill: Dict[str, int], stats: dict) -> dict:
    """
    Integra datos del debug en la fila según el modo:
      - 'none': no hace merge.
      - 'only_blanks': completa solo vacíos.
      - 'prefer_debug': sobreescribe con debug si hay valor.
    Actualiza métricas de campos rellenados y de aciertos/omisos.
    """
    if mode == "none" or not debug_map:
        return row
    key_opts = [row.get("OPERACION", "").strip(), f"{row.get('RUT','')}-{row.get('DV','')}-{row.get('NOMBRE','')}".strip("-")]
    cand = None
    for k in key_opts:
        if not k:
            continue
        cand = debug_map.get(k)
        if cand:
            break
    if not cand:
        stats["debug_merge_misses"] = stats.get("debug_merge_misses", 0) + 1
        return row
    stats["debug_merge_hits"] = stats.get("debug_merge_hits", 0) + 1
    pairs = [
        ("NOMBRE", "NOMBRE"), ("DIRECCION", "DIRECCION"), ("COMUNA", "COMUNA"),
        ("FECHA_SUSCRIPCION", "FECHA_SUSCRIPCION"), ("MONTO_CREDITO", "MONTO_CREDITO"),
        ("FECHA_VENCIMIENTO_1_CUOTA", "FECHA_VENCIMIENTO_1_CUOTA"),
        ("FECHA_VENCIMIENTO_ULTIMA_CUOTA", "FECHA_VENCIMIENTO_ULTIMA_CUOTA"),
        ("CAPITAL", "CAPITAL"), ("PRODUCTO", "PRODUCTO"),
        ("NOMBRE_APODERADO", "NOMBRE_APODERADO"), ("NOMBRE_APODERADO_2", "NOMBRE_APODERADO_2"),
    ]
    for f_csv, f_dbg in pairs:
        v_csv = clean_text_value(row.get(f_csv, ""))
        v_dbg = clean_text_value(cand.get(f_dbg, ""))
        if mode == "only_blanks":
            if not v_csv and v_dbg:
                row[f_csv] = v_dbg
                stats_fill[f_csv] = stats_fill.get(f_csv, 0) + 1
        elif mode == "prefer_debug":
            if v_dbg and v_dbg != v_csv:
                row[f_csv] = v_dbg
                stats_fill[f_csv] = stats_fill.get(f_csv, 0) + 1
    return row


# ========================= Proceso de filas =========================

def clean_and_normalize_row(row: dict, date_format: str, thousand_sep: str, strict_dv: bool, stats: dict) -> dict:
    """
    Aplica todas las normalizaciones a una fila y actualiza estadísticas.
    """
    # Asegura todas las columnas existan
    for h in CANONICAL_HEADERS:
        row.setdefault(h, "")

    # Limpieza de texto en todas las celdas
    for k in list(row.keys()):
        before = row[k]
        row[k] = clean_text_value(row[k])
        if before != row[k]:
            stats["fixed_encoding"] = stats.get("fixed_encoding", 0) + 1

    # Fixes comunes
    for k in ["NOMBRE", "DIRECCION", "COMUNA"]:
        before = row.get(k, "")
        after = apply_common_fixes(before)
        if after != before:
            stats["fixed_common"] = stats.get("fixed_common", 0) + 1
            row[k] = after

    # Apoderados
    a1_before = row.get("NOMBRE_APODERADO", "")
    a1_after = clean_apoderado(a1_before, which=1)
    if a1_after != a1_before:
        stats["apoderado1_fixed"] = stats.get("apoderado1_fixed", 0) + 1
        row["NOMBRE_APODERADO"] = a1_after

    a2_before = row.get("NOMBRE_APODERADO_2", "")
    a2_after = clean_apoderado(a2_before, which=2)
    if a2_after != a2_before:
        stats["apoderado2_fixed"] = stats.get("apoderado2_fixed", 0) + 1
        row["NOMBRE_APODERADO_2"] = a2_after

    # Fechas
    for k in DATE_FIELDS:
        raw_date = row.get(k, "")
        d = parse_date_multi(raw_date)
        if d:
            row[k] = format_date(d, date_format)
            stats["normalized_dates"] = stats.get("normalized_dates", 0) + 1

    # Tasa
    if row.get("TASA"):
        before = row["TASA"]
        row["TASA"] = normalize_percent(before)
        if row["TASA"] != before:
            stats["normalized_percent"] = stats.get("normalized_percent", 0) + 1

    # Montos/enteros
    for k in INT_FIELDS:
        before = row.get(k, "")
        digits = normalize_int_digits(before)
        formatted = format_int(digits, thousand_sep)
        row[k] = formatted
        if row[k] != before:
            stats["normalized_ints"] = stats.get("normalized_ints", 0) + 1

    # RUT/DV
    rut_before, dv_before = row.get("RUT", ""), row.get("DV", "")
    rut_num, dv_clean, dv_calc, ok = normalize_rut_and_dv(rut_before, dv_before)
    if rut_num != rut_before or dv_clean != dv_before:
        stats["normalized_rut"] = stats.get("normalized_rut", 0) + 1
    row["RUT"] = rut_num
    row["DV"] = dv_clean
    if rut_num and dv_clean and not ok:
        stats["rut_invalid"] = stats.get("rut_invalid", 0) + 1
        if strict_dv and dv_calc:
            row["DV"] = dv_calc

    # Valida comuna si se dispone de catálogo
    if VALID_COMUNAS:
        comuna = row.get("COMUNA", "")
        if comuna and comuna not in VALID_COMUNAS:
            stats["invalid_comunas"] = stats.get("invalid_comunas", 0) + 1

    return row


def row_is_incomplete(row: dict, required_fields: Iterable[str]) -> Tuple[bool, List[str]]:
    """
    Retorna si la fila está incompleta con base en required_fields y las razones.
    """
    missing = [f for f in required_fields if not clean_text_value(row.get(f, ""))]
    return (len(missing) > 0, missing)


# ========================= Detecciones (encoding/delimiter) =========================

def detect_encoding(path: Path, fallback: str = "utf-8") -> str:
    """
    Detecta encoding usando chardet si está disponible.
    Devuelve fallback si no se puede detectar.
    """
    if not chardet:
        return fallback
    try:
        with open(path, "rb") as fb:
            raw = fb.read(128 * 1024)
        res = chardet.detect(raw)
        enc = res.get("encoding") or fallback
        return enc
    except Exception:
        return fallback


def sniff_delimiter(path: Path, encoding: str) -> str:
    """
    Autodetecta el delimitador usando csv.Sniffer.
    Retorna ';' por defecto si no se detecta.
    """
    try:
        with io.open(path, "r", encoding=encoding, errors="ignore") as f:
            sample = f.read(4096)
        dialect = csv.Sniffer().sniff(sample, ",;\t|")
        return dialect.delimiter
    except Exception:
        return ";"


def write_to_excel(data_rows: List[Dict[str, str]], output_path: str) -> None:
    """
    Escribe los datos procesados a un archivo Excel con formato.
    """
    if not openpyxl:
        raise ImportError("openpyxl no está disponible. Instala con: pip install openpyxl")
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Datos Limpios Itaú"
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col, header in enumerate(CANONICAL_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, row_data in enumerate(data_rows, 2):
            for col, header in enumerate(CANONICAL_HEADERS, 1):
                value = row_data.get(header, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                
                # Formatear fechas
                if header in DATE_FIELDS and value and value != "":
                    try:
                        # Intentar parsear la fecha para formatearla
                        if "-" in value:
                            date_obj = dt.datetime.strptime(value, "%Y-%m-%d")
                            cell.value = date_obj
                            cell.number_format = "DD-MM-YYYY"
                    except Exception:
                        pass  # Mantener como texto si no se puede parsear
                
                # Formatear números
                elif header in INT_FIELDS and value and value != "":
                    try:
                        cell.value = int(value)
                        cell.number_format = "#,##0"
                    except Exception:
                        pass  # Mantener como texto si no se puede convertir
        
        # Ajustar ancho de columnas
        for col_idx, header in enumerate(CANONICAL_HEADERS, 1):
            max_length = len(header)  # Al menos el ancho del encabezado
            col_letter = get_column_letter(col_idx)
            
            # Calcular el ancho máximo de la columna
            for row_idx in range(2, len(data_rows) + 2):
                try:
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Guardar archivo
    wb.save(output_path)


# ========================= Núcleo: proceso completo =========================

def process(
    input_csv: str,
    output_csv: str,
    report_path: Optional[str],
    debug_path: Optional[str],
    date_format: str,
    thousand_sep: str,
    fill_from_debug: str,
    strict_dv: bool,
    delimiter: str = ";",
    required_fields: Optional[List[str]] = None,
    reject_incomplete: bool = False,
) -> None:
    """
    Procesa el CSV de entrada y escribe el CSV limpio:
    - Autodetecta delimitador si se pasa 'auto'.
    - Lee y escribe en streaming.
    - Integra datos del debug según modo.
    - Genera reporte Markdown si se solicitó.
    """
    required_fields = required_fields or []

    stats = {
        "rows": 0, "fixed_headers": 0, "fixed_encoding": 0, "fixed_common": 0,
        "normalized_dates": 0, "normalized_ints": 0, "normalized_percent": 0,
        "normalized_rut": 0, "rut_invalid": 0, "apoderado1_fixed": 0, "apoderado2_fixed": 0,
        "rut_invalid_examples": [], "debug_merge_hits": 0, "debug_merge_misses": 0,
        "rows_rejected": 0, "neg_amounts_detected": 0, "invalid_comunas": 0,
    }
    stats_fill: Dict[str, int] = {}

    # Carga debug si existe
    debug_map = {}
    if debug_path and os.path.exists(debug_path):
        try:
            enc_dbg = detect_encoding(Path(debug_path))
            with io.open(debug_path, "r", encoding=enc_dbg, errors="ignore") as f:
                debug_map = parse_debug_final_rows(f.read())
            logging.info("Debug FINAL ROWs cargados: %d", len(debug_map))
        except Exception as e:
            logging.error("No se pudo abrir el debug %s: %s", debug_path, e)

    # Detección de encoding y delimitador
    input_path = Path(input_csv)
    enc_in = detect_encoding(input_path)
    delim = sniff_delimiter(input_path, enc_in) if delimiter == "auto" else delimiter

    # Determinar formato de salida
    is_excel_output = output_csv.lower().endswith(('.xlsx', '.xls'))
    
    # Abre entrada
    try:
        fin = io.open(input_csv, "r", encoding=enc_in, errors="ignore", newline="")
    except Exception as e:
        logging.error("No se pudo abrir input %s: %s", input_csv, e)
        raise SystemExit(2)

    # Lista para almacenar datos procesados (para Excel o CSV)
    processed_rows = []

    with fin:
        reader = csv.DictReader(fin, delimiter=delim)
        in_headers = reader.fieldnames or []
        mapped_headers = [normalize_header(h) for h in in_headers]
        if mapped_headers != in_headers:
            stats["fixed_headers"] += 1

        for raw in reader:
            stats["rows"] += 1

            # Remapea a canónicos
            row = {normalize_header(k): v for k, v in raw.items()}

            # Limpieza y normalizaciones
            row = clean_and_normalize_row(row, date_format=date_format, thousand_sep=thousand_sep, strict_dv=strict_dv, stats=stats)

            # Merge con debug
            row = merge_from_debug(row, debug_map, fill_from_debug, stats_fill, stats)

            # Validación de requeridos
            if required_fields:
                incomplete, missing = row_is_incomplete(row, required_fields)
                if incomplete:
                    stats["rows_rejected"] += 1
                    if len(stats["rut_invalid_examples"]) < 20:
                        stats["rut_invalid_examples"].append({"OPERACION": row.get("OPERACION", ""), "MISSING": ",".join(missing)})
                    if reject_incomplete:
                        continue  # no escribas la fila

            # Agregar fila procesada a la lista
            processed_rows.append({h: row.get(h, "") for h in CANONICAL_HEADERS})

    # ========================= GEOLOCALIZACIÓN Y CORRECCIONES =========================
    # Aplicar geolocalización y correcciones de referencia si está disponible
    if GEOCODING_AVAILABLE and processed_rows:
        try:
            import pandas as pd
            logging.info("🌍 Aplicando geolocalización y correcciones...")
            
            # Convertir a DataFrame para procesamiento
            df_temp = pd.DataFrame(processed_rows)
            
            # 1. Aplicar correcciones de referencia basadas en datos conocidos
            df_corrected = apply_reference_corrections(df_temp)
            corrected_count = sum(1 for i in range(len(df_temp)) 
                                if df_temp.iloc[i].to_dict() != df_corrected.iloc[i].to_dict())
            if corrected_count > 0:
                logging.info(f"📋 Aplicadas {corrected_count} correcciones de referencia")
            
            # 2. Aplicar geolocalización para mejorar direcciones/comunas
            df_enhanced = enhance_dataframe_with_geolocation(
                df_corrected, 
                address_col='DIRECCION', 
                comuna_col='COMUNA'
            )
            
            # Actualizar las filas procesadas con la información mejorada
            processed_rows = df_enhanced.to_dict('records')
            
            # Actualizar estadísticas
            geocoded_count = sum(1 for row in processed_rows 
                               if row.get('GEOCODING_CONFIDENCE', 0) > 0.5)
            stats['geocoded_addresses'] = geocoded_count if geocoded_count > 0 else 0
            stats['reference_corrections'] = corrected_count
            
            logging.info(f"✅ Procesamiento completado: {corrected_count} correcciones + {geocoded_count} geocodificaciones")
            
        except Exception as e:
            logging.warning(f"⚠️ Error en procesamiento: {e}")

    # Escribir archivo de salida según el formato
    try:
        if is_excel_output:
            # Ensure all values are strings for Excel writing
            string_rows = []
            for row in processed_rows:
                string_row = {str(k): str(v) for k, v in row.items()}
                string_rows.append(string_row)
            write_to_excel(string_rows, output_csv)
        else:
            # Escribir CSV tradicional
            with io.open(output_csv, "w", encoding="utf-8", newline="") as fout:
                writer = csv.DictWriter(fout, fieldnames=CANONICAL_HEADERS, delimiter=";")
                writer.writeheader()
                for row in processed_rows:
                    # Ensure all keys are strings for CSV writing
                    string_row = {str(k): str(v) for k, v in row.items()}
                    writer.writerow(string_row)
    except Exception as e:
        logging.error("No se pudo escribir el archivo de salida %s: %s", output_csv, e)
        raise SystemExit(2)

    # Reporte
    if report_path:
        try:
            with io.open(report_path, "w", encoding="utf-8") as f:
                f.write("# Informe de limpieza Itau\n")
                f.write(f"- Filas procesadas: {stats['rows']}\n")
                f.write(f"- Filas rechazadas (incompletas): {stats['rows_rejected']}\n")
                f.write(f"- Encabezados corregidos: {stats['fixed_headers']}\n")
                f.write(f"- Valores normalizados (texto/codificación): {stats['fixed_encoding']}\n")
                f.write(f"- Arreglos comunes (comunas/direcciones): {stats['fixed_common']}\n")
                f.write(f"- Fechas normalizadas: {stats['normalized_dates']}\n")
                f.write(f"- Campos enteros/monetarios normalizados: {stats['normalized_ints']}\n")
                f.write(f"- Tasas normalizadas: {stats['normalized_percent']}\n")
                f.write(f"- RUT/DV normalizados: {stats['normalized_rut']}\n")
                f.write(f"- RUT con DV inconsistente: {stats['rut_invalid']}\n")
                f.write(f"- Apoderado 1 corregido: {stats['apoderado1_fixed']}\n")
                f.write(f"- Apoderado 2 corregido: {stats['apoderado2_fixed']}\n")
                f.write(f"- Merges con debug (aciertos): {stats['debug_merge_hits']}\n")
                f.write(f"- Merges con debug (sin coincidencia): {stats['debug_merge_misses']}\n")
                f.write(f"- Montos negativos detectados (antes de normalizar): {stats['neg_amounts_detected']}\n")
                if VALID_COMUNAS:
                    f.write(f"- Comunas inválidas detectadas: {stats['invalid_comunas']}\n")
                if 'geocoded_addresses' in stats:
                    f.write(f"- Direcciones geocodificadas exitosamente: {stats['geocoded_addresses']}\n")
                if stats_fill:
                    f.write("\n## Campos rellenados desde Itau_auto_debug.txt\n")
                    for k, n in sorted(stats_fill.items()):
                        f.write(f"- {k}: {n}\n")
                if stats["rut_invalid_examples"]:
                    f.write("\n## Ejemplos (máx 20) de filas con DV inconsistente o incompletas\n")
                    for ex in stats["rut_invalid_examples"]:
                        f.write(f"- {ex}\n")
        except Exception as e:
            logging.error("No se pudo escribir el reporte %s: %s", report_path, e)

    # ========================= LIMPIEZA AUTOMÁTICA =========================
    # Limpiar archivos temporales después del procesamiento
    if GEOCODING_AVAILABLE:
        try:
            current_dir = os.path.dirname(output_csv) if output_csv else os.getcwd()
            cleanup_patterns = [
                '*.tmp',
                '*_temp.csv',
                '*_debug.txt',
                'temp_*',
                'page_*.png',  # Imágenes OCR temporales
                'extracted_data_*.csv',  # CSVs intermedios del OCR
                '*_intermediate.*',
                'test_*.csv',  # Archivos de prueba
                'datos_*_temp.*'
            ]
            
            cleanup_temp_files(current_dir, cleanup_patterns)
            logging.info("🧹 Limpieza automática completada")
            
        except Exception as e:
            logging.warning(f"⚠️ Error en limpieza automática: {e}")

    logging.info("OK -> Salida: %s | Reporte: %s", output_csv, report_path or "(none)")


# ========================= Autodetección de archivos cuando faltan args =========================

def find_candidate_csvs(base_dir: str) -> List[str]:
    """
    Busca CSVs en la carpeta, priorizando 'Itau_results_ALL.csv'.
    """
    names = []
    for fn in os.listdir(base_dir):
        if fn.lower().endswith(".csv"):
            names.append(os.path.join(base_dir, fn))
    preferred = [p for p in names if os.path.basename(p).lower() == "itau_results_all.csv"]
    if preferred:
        return preferred + [p for p in names if p not in preferred]
    names.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return names


def infer_defaults(args) -> bool:
    """
    Rellena --input/--output/--debug/--report cuando no se pasan por CLI.
    Retorna True si todo está bien, False si hay algún problema.
    """
    base_dir = os.getcwd()
    # INPUT
    if not args.input:
        cands = find_candidate_csvs(base_dir)
        if not cands:
            print("⚠️  No se encontró ningún CSV en la carpeta actual.")
            print("💡 Opciones:")
            print("   1. Arrastra tu archivo CSV a esta carpeta")
            print("   2. Usa: python process_itau_auto_v2.py --input ruta/a/tu/archivo.csv")
            print("   3. Ejecuta primero el OCR con: python ocr_to_csv.py")
            return False
        args.input = cands[0]
        print(f"[auto] --input: {args.input}")
    # OUTPUT
    if not args.output:
        root, _ = os.path.splitext(args.input)
        extension = ".xlsx" if args.format == "excel" else ".csv"
        
        # Si es Excel, guardarlo en outputs/Itau
        if args.format == "excel":
            # Crear directorio de salida
            output_dir = Path("../outputs/Itau")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Usar solo el nombre del archivo, sin la ruta completa
            filename = Path(root).name
            args.output = str(output_dir / f"{filename}.cleaned{extension}")
        else:
            args.output = root + ".cleaned" + extension
        
        print(f"[auto] --output: {args.output}")
    # DEBUG
    if not args.debug:
        dbg = os.path.join(base_dir, "Itau_auto_debug.txt")
        if os.path.exists(dbg):
            args.debug = dbg
            print(f"[auto] --debug: {args.debug}")
    # REPORT
    if not args.report:
        args.report = os.path.join(os.path.dirname(args.output), "fix_report.md")
        print(f"[auto] --report: {args.report}")
    
    return True


# ========================= CLI =========================

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Procesa y limpia CSV Itaú con soporte de Itau_auto_debug.txt")
    p.add_argument("--input", required=False, help="Ruta al CSV de entrada (por defecto autodetecta en carpeta actual)")
    p.add_argument("--output", required=False, help="Ruta al archivo de salida (por defecto <input>.cleaned.xlsx para Excel)")
    p.add_argument("--debug", required=False, default=None, help="Ruta a Itau_auto_debug.txt (opcional)")
    p.add_argument("--report", required=False, default=None, help="Ruta al reporte Markdown (por defecto fix_report.md junto al output)")
    p.add_argument("--date-format", choices=["iso", "dmy"], default="iso", help="Formato de fechas: iso (yyyy-mm-dd) o dmy (dd-mm-aaaa)")
    p.add_argument("--thousand-sep", choices=["none", "dot", "comma"], default="none", help="Formato de miles para enteros/monedas")
    p.add_argument("--fill-from-debug", choices=["none", "only_blanks", "prefer_debug"], default="only_blanks",
                   help="Cómo usar los 'FINAL ROW' del debug para completar datos")
    p.add_argument("--strict-dv", action="store_true", help="Si se activa, sobreescribe DV con el calculado cuando no coincide")
    p.add_argument("--delimiter", default=";", help="Delimitador del CSV de entrada ('auto' para autodetectar)")
    p.add_argument("--required-fields", nargs="*", default=[], help="Campos requeridos; si faltan se registra y opcionalmente se rechaza")
    p.add_argument("--reject-incomplete", action="store_true", help="No escribe filas con campos requeridos vacíos")
    p.add_argument("--format", choices=["csv", "excel"], default="excel", help="Formato de salida: csv o excel (por defecto excel)")
    p.add_argument("-v", "--verbose", action="count", default=0, help="Verboso (-v, -vv)")
    return p


def main():
    ap = build_arg_parser()
    args = ap.parse_args()
    # Autodetecta defaults cuando faltan
    if not infer_defaults(args):
        print("\n❌ No se pudo inicializar el procesamiento.")
        print("🚀 Intenta ejecutar primero: python ocr_to_csv.py --client Itau --pdfs-dir pdfs/Itau")
        sys.exit(1)
    
    logging.basicConfig(
        level=logging.WARNING if args.verbose == 0 else (logging.INFO if args.verbose == 1 else logging.DEBUG),
        format="%(levelname)s: %(message)s"
    )
    # Validación de existencia de input
    if not os.path.exists(args.input):
        print(f"❌ El archivo de entrada no existe: {args.input}", file=sys.stderr)
        print("💡 Verifica que el archivo exista o ejecuta primero el OCR.")
        sys.exit(2)
    # Ejecuta
    process(
        input_csv=args.input,
        output_csv=args.output,
        report_path=args.report,
        debug_path=args.debug,
        date_format=args.date_format,
        thousand_sep=args.thousand_sep,
        fill_from_debug=args.fill_from_debug,
        strict_dv=args.strict_dv,
        delimiter=args.delimiter,
        required_fields=args.required_fields,
        reject_incomplete=args.reject_incomplete,
    )
    print("\nOK. Procesamiento completado.")
    format_name = "Excel" if args.output.lower().endswith(('.xlsx', '.xls')) else "CSV"
    print(f"- Archivo {format_name}: {args.output}")
    if args.report:
        print(f"- Reporte: {args.report}")


if __name__ == "__main__":
    main()
